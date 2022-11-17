
import openpyxl

path=r"D:\Naushad\ReimbursementsDefectsNGCopy.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active     #workbook.get_sheet_by_name("sheet1")

rows=sheet.max_row
cols=sheet.max_column

print(rows)   #19
print(cols)   #5

for r in range(1,rows+1):  #range function considers n-1 so we have used rows+1 here
    for c in range(1,cols+1):
        print(sheet.cell(row=r, column=c).value, end="    ")
    print()










